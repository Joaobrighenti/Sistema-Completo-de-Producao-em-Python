import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import io
import base64

def format_number(val):
    """Formata números para exibição"""
    if val is not None and not pd.isna(val):
        return f"{val:,.0f}".replace(',', '.')
    return "0"

def get_status_color(pedido_mensal, imp_qtd, feito_qtd):
    """Determina a cor do status baseado nas quantidades"""
    if pd.notna(pedido_mensal) and pedido_mensal > 0:
        soma_imp_feito = imp_qtd + feito_qtd
        ratio = soma_imp_feito / pedido_mensal
        if ratio >= 2: return 'd4edda'  # Verde claro
        if ratio >= 1: return 'fff3cd'  # Amarelo claro
        return 'f8d7da'  # Vermelho claro
    return 'f0f0f0'  # Cinza

def get_indicator_color(tipo_trabalho):
    """Retorna a cor do indicador baseado no tipo de trabalho"""
    colors = {
        'Semanal': '28a745',  # Verde
        'Quinzenal': 'ffc107',  # Amarelo
        'Mensal': 'dc3545'  # Vermelho
    }
    return colors.get(tipo_trabalho, '6c757d')  # Cinza padrão

def create_excel_entregas_programadas(daily_data, center_day_index):
    """Cria um arquivo Excel com a estrutura das entregas programadas"""
    wb = Workbook()
    
    # Verifica se há dados para processar
    if not daily_data:
        # Usa a planilha padrão se não há dados
        ws = wb.active
        ws.title = "Sem Dados"
        ws['A1'] = "Nenhum dado disponível para exportação."
        return wb
    
    # Remove a planilha padrão apenas se há dados
    wb.remove(wb.active)
    
    # Dias da semana
    weekdays = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado"]
    
    # Cores para formatação
    header_fill = PatternFill(start_color="818a89", end_color="818a89", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Determina quais dias mostrar (dia anterior, atual e próximo)
    num_days = len(weekdays)
    indices_to_show = [
        (center_day_index - 1 + num_days) % num_days,
        center_day_index,
        (center_day_index + 1) % num_days
    ]
    
    planilhas_criadas = False
    for index in indices_to_show:
        day_name = weekdays[index]
        # Converte a chave de string para inteiro se necessário
        day_key = str(index) if str(index) in daily_data else index
        day_data = daily_data.get(day_key, {})
        
        # Cria uma nova planilha para cada dia
        ws = wb.create_sheet(title=day_name)
        planilhas_criadas = True
        
        # Se não há dados para este dia, mostra mensagem e continua
        if not day_data:
            ws['A1'] = f"Nenhuma entrega programada para {day_name}."
            ws['A1'].font = Font(italic=True, color="666666")
            continue
        
        # Configura larguras das colunas
        ws.column_dimensions['A'].width = 40  # Produto
        ws.column_dimensions['B'].width = 8   # Indicador
        ws.column_dimensions['C'].width = 12  # Pedido
        ws.column_dimensions['D'].width = 12  # Impresso
        ws.column_dimensions['E'].width = 12  # Feito
        
        current_row = 1
        
        for client_name, products in day_data.items():
            # Cabeçalho do cliente
            ws[f'A{current_row}'] = client_name
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            ws[f'A{current_row}'].fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")
            ws.merge_cells(f'A{current_row}:E{current_row}')
            current_row += 1
            
            # Cabeçalho das colunas
            headers = ['Produto', 'Ind.', 'Pedido', 'Impresso', 'Feito']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            current_row += 1
            
            # Dados dos produtos
            for product_info in products:
                product_name = product_info.get('produto_nome', '')
                tipo_trabalho = product_info.get('tipo_trabalho', '')
                pedido_mensal = product_info.get('pedido_mensal', 0)
                imp_qtd = product_info.get('imp_qtd', 0)
                feito_qtd = product_info.get('feito_qtd', 0)
                
                # Determina cor de fundo do pedido
                bg_color = get_status_color(pedido_mensal, imp_qtd, feito_qtd)
                
                # Produto
                ws.cell(row=current_row, column=1, value=product_name)
                ws.cell(row=current_row, column=1).border = border
                
                # Indicador
                if tipo_trabalho:
                    indicator = tipo_trabalho[0] if tipo_trabalho else ''
                    indicator_cell = ws.cell(row=current_row, column=2, value=indicator)
                    indicator_cell.alignment = Alignment(horizontal='center', vertical='center')
                    indicator_cell.font = Font(bold=True, color="FFFFFF")
                    indicator_cell.fill = PatternFill(start_color=get_indicator_color(tipo_trabalho), 
                                                     end_color=get_indicator_color(tipo_trabalho), 
                                                     fill_type="solid")
                    indicator_cell.border = border
                
                # Pedido
                pedido_cell = ws.cell(row=current_row, column=3, value=format_number(pedido_mensal))
                pedido_cell.alignment = Alignment(horizontal='center', vertical='center')
                pedido_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                pedido_cell.border = border
                
                # Impresso
                imp_cell = ws.cell(row=current_row, column=4, value=format_number(imp_qtd))
                imp_cell.alignment = Alignment(horizontal='center', vertical='center')
                imp_cell.border = border
                
                # Feito
                feito_cell = ws.cell(row=current_row, column=5, value=format_number(feito_qtd))
                feito_cell.alignment = Alignment(horizontal='center', vertical='center')
                feito_cell.border = border
                
                current_row += 1
            
            # Espaço entre clientes
            current_row += 1
    
    # Verifica se pelo menos uma planilha foi criada
    if not planilhas_criadas:
        ws = wb.create_sheet(title="Sem Dados")
        ws['A1'] = "Nenhum dado disponível para exportação."
    
    return wb

def create_excel_estoque_empurrado(df_final):
    """Cria um arquivo Excel com a estrutura do estoque empurrado"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque Empurrado"
    
    # Verifica se há dados
    if df_final.empty:
        ws['A1'] = "Nenhum dado disponível para exportação."
        return wb
    
    # Cabeçalhos
    headers = ["ID PRODUTO", "PRODUTO", "PREVISÃO MENSAL", "ESTOQUE", "COBERTURA PROJEÇÃO"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Dados
    for row_idx, (_, row) in enumerate(df_final.iterrows(), 2):
        for col_idx, value in enumerate(row[['produto_id', 'nome', 'pedido_mensal', 'saldo_em_estoque', 'seta_por_pedido']], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Formatação especial para a coluna de cobertura
            if col_idx == 5:  # Cobertura Projeção
                if '↑' in str(value):
                    cell.font = Font(color="28a745")  # Verde
                elif '!' in str(value):
                    cell.font = Font(color="ffc107")  # Amarelo
                elif '↓' in str(value):
                    cell.font = Font(color="dc3545")  # Vermelho
    
    # Ajusta larguras das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    
    return wb

def excel_to_base64(wb):
    """Converte o workbook para base64 para download"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Codifica em base64
    encoded = base64.b64encode(output.read()).decode()
    return encoded

def generate_excel_download(daily_data, center_day_index, view_type='programadas', df_empurrado=None):
    """Gera o arquivo Excel para download baseado no tipo de visualização"""
    if view_type == 'programadas':
        wb = create_excel_entregas_programadas(daily_data, center_day_index)
    elif view_type == 'empurrado':
        if df_empurrado is not None and not df_empurrado.empty:
            wb = create_excel_estoque_empurrado(df_empurrado)
        else:
            # Cria um workbook vazio se não há dados
            wb = Workbook()
            ws = wb.active
            ws.title = "Estoque Empurrado"
            ws['A1'] = "Nenhum dado disponível para exportação."
    else:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Tipo de visualização não suportado."
    
    return excel_to_base64(wb)
